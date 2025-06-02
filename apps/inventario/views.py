from django.shortcuts import render, redirect, get_object_or_404
from .forms import InventarioForm
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from apps.inventario.models import m_inventario, MovimientoInventario, Categoria
from .forms import ExcelUploadForm
import openpyxl
import os

# Create your views here.
from django.db.models import Q

def p_productos(request):
    # Valores iniciales
    stock_minimo = 5  # Puedes hacer esto configurable como vimos antes
    productos = m_inventario.objects.all()
    # Filtros
    query = request.GET.get('query')  # Para búsqueda general
    query2 = request.GET.get('busca_producto')  # Para búsqueda específica
    solo_minimos = request.GET.get('solo_minimos') == '1'
    # Aplicar filtros en el orden correcto
    if query:
        productos = productos.filter(
            Q(nombre_p__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(categoria__categoria__icontains=query)
        )
    if query2:
        productos = productos.filter(
            Q(nombre_p__icontains=query2)  # Cambiado a nombre_p para coincidir con tu modelo
        )
    if solo_minimos:
        productos = productos.filter(stock__lte=stock_minimo)
    # Contador de productos en mínimo
    total_minimos = productos.filter(stock__lte=stock_minimo).count()

    context = {
        'productos': productos,
        'stock_minimo': stock_minimo,
        'query': query or query2,  # Para mostrar en el input de búsqueda
        'total_minimos': total_minimos,
        'solo_minimos': solo_minimos,
    }
    return render(request, 'inventario/inventario.html', context)


def agregar_producto(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST) #Para guardar un formulario
        if form.is_valid():
            form.save()
            messages.success(request, 'producto|agregado|Producto agregado correctamente')
            return redirect('inventario') #Redireccionamiento
    else:
        form = InventarioForm()

    return render(request, 'inventario/agregar_producto.html',  {'form': form})

def ver_producto(request, id_producto):
    producto = get_object_or_404(m_inventario, pk=id_producto) #Para pasar la id
    return render(request, 'inventario/ver_producto.html', {'producto': producto})

def editar_producto(request, id_producto):
    producto = get_object_or_404(m_inventario, pk=id_producto)
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'prodcuto|actualizado|Prodcuto actualizado correctamente')
            return redirect('inventario')
    else:
        form = InventarioForm(instance=producto)
    return render(request, 'inventario/editar_producto.html', {'form': form, 'producto': producto})

@require_POST
def eliminar_producto(request):
    id_producto = request.POST.get('id_producto')
    producto = get_object_or_404(m_inventario, pk=id_producto)
    producto.delete()
    messages.success(request, 'producto|eliminado|Producto eliminado correctamente')
    return redirect('inventario')

def actualizar_stock(request, id_producto):
    producto = get_object_or_404(m_inventario, pk=id_producto)
    # stock_anterior = producto.stock

    if request.method == 'POST':
        try:
            cantidad_agregada = int(request.POST.get('stock'))
            if cantidad_agregada <= 0:
                raise ValueError("¡Cantidad invalida! Intentalo de nuevo")
            producto.stock += cantidad_agregada
            producto.save()
            MovimientoInventario.objects.create(
                    producto=producto,
                    tipo='entrada',
                    cantidad=cantidad_agregada,
                    fecha=timezone.now())
            messages.success(request, 'prodcuto|actualizado|Stock agregado correctamente')
            return redirect('inventario')
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')

    return render(request, 'inventario/actualizar_stock.html', {'producto': producto})

def cargar_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']

        # Configuración de mapeo basada en tu modelo
        FIELD_CONFIG = {
            'nombre': {
                'model_field': 'nombre_p',
                'type': str,
                'required': True,
                'default': None,
                'clean': lambda x: x.strip() if x else None
            },
            'descripcion': {
                'model_field': 'descripcion',
                'type': str,
                'required': False,
                'default': None,
                'clean': lambda x: x.strip() if x else None
            },
            'categoria': {
                'model_field': 'categoria',
                'type': 'fk',
                'required': False,
                'default': None,
                'clean': lambda x: x.strip() if x else None
            },
            'stock': {
                'model_field': 'stock',
                'type': int,
                'required': False,
                'default': None,
                'clean': lambda x: int(float(x)) if x not in (None, '', ' ') else None
            },
            'precio': {
                'model_field': 'precio',
                'type': float,
                'required': False,
                'default': 0.00,
                'clean': lambda x: round(float(x), 2) if x not in (None, '', ' ') else 0.00
            }
        }

        try:
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active

            # Normalizar encabezados (minúsculas, sin espacios)
            encabezados = [
                cell.value.lower().replace(' ', '').strip() 
                for cell in next(hoja.iter_rows(min_row=1, max_row=1))
                if cell.value
            ]

            # Crear mapeo de columnas
            column_map = {}
            for excel_col, config in FIELD_CONFIG.items():
                if excel_col in encabezados:
                    column_map[excel_col] = {
                        'index': encabezados.index(excel_col),
                        'config': config
                    }

            registros_creados = 0
            errores = []

            for row_num, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                if not any(fila):  # Saltar filas vacías
                    continue
                try:
                    data = {}
                    # Procesar cada campo encontrado en el Excel
                    for excel_col, mapping in column_map.items():
                        raw_value = fila[mapping['index']]
                        config = mapping['config']
                        try:
                            # Limpiar y convertir el valor
                            cleaned_value = config['clean'](raw_value)
                            # Manejo especial para FK (categoría)
                            if config['type'] == 'fk' and excel_col == 'categoria' and cleaned_value:
                                categoria, _ = Categoria.objects.get_or_create(categoria=cleaned_value)
                                data[config['model_field']] = categoria
                            else:
                                data[config['model_field']] = cleaned_value if cleaned_value is not None else config['default']
                        except Exception as e:
                            raise ValueError(f"Error en columna '{excel_col}': {str(e)}")

                    # Validar campo requerido
                    if not data.get('nombre_p'):
                        raise ValueError("El nombre del producto es requerido")
                    # Crear registro
                    m_inventario.objects.create(**data)
                    registros_creados += 1
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
                    continue

            # Reporte de resultados
            if registros_creados > 0:
                if errores:
                    msg += f" | {len(errores)} errores"
                messages.success(request, 'producto|agregado|Importación exitosa: Productos agregados correctamente')
            else:
                messages.warning(request, "producto|error|Ningún registro fue importado")
            if errores:
                error_samples = "\n".join(errores[:5])
                messages.info(request, f"producto|error|Errores encontrados:\n{error_samples}")
        except Exception as e:
            messages.error(request, 'producto|error|Error al procesar el archivo')
            # Log detallado para debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error en importación Excel: {str(e)}", exc_info=True)

    return redirect('inventario')